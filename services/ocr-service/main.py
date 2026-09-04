import os
import io
import csv
import json
import logging
import concurrent.futures
from typing import Optional
from fastapi import FastAPI, HTTPException, status
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
import boto3
from botocore.client import Config
from PIL import Image
import pytesseract
import pdf2image
import pypdf
import openpyxl

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("ocr-service")

app = FastAPI(
    title="MineIQ OCR & Parsing Service",
    description="Stateless document parser extracting text from PDFs, images, and spreadsheets for CIL/CMPDI platform."
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configuration from environment
MINIO_ENDPOINT = os.getenv("MINIO_ENDPOINT", "http://minio:9000")
MINIO_ACCESS_KEY = os.getenv("MINIO_ROOT_USER", os.getenv("MINIO_ACCESS_KEY", "minioadmin"))
MINIO_SECRET_KEY = os.getenv("MINIO_ROOT_PASSWORD", os.getenv("MINIO_SECRET_KEY", "minioadmin"))
MINIO_BUCKET = os.getenv("MINIO_BUCKET", "raw-files")
OCR_TIMEOUT_SECONDS = float(os.getenv("OCR_TIMEOUT_SECONDS", "60.0"))


def get_s3_client():
    return boto3.client(
        's3',
        endpoint_url=MINIO_ENDPOINT,
        aws_access_key_id=MINIO_ACCESS_KEY,
        aws_secret_access_key=MINIO_SECRET_KEY,
        config=Config(signature_version='s3v4'),
        region_name='us-east-1'
    )


import re

class ExtractRequest(BaseModel):
    document_id: str
    s3_key: str
    source_type: str


class ExtractResponse(BaseModel):
    document_id: str
    extracted_text: str
    method_used: str
    page_count: Optional[int] = None
    char_count: int
    structured_data: Optional[dict] = None
    structured_records: Optional[list] = None
    error: Optional[str] = None


SUBSIDIARY_RE = re.compile(r'\b(ECL|BCCL|CCL|WCL|SECL|MCL|NCL|CMPDI)\b', re.IGNORECASE)

# Full legal names -> subsidiary code, so documents that spell out the company
# (as official reports usually do) are still tagged. Ordered longest-first at
# match time; SECL before ECL/CCL is handled by explicit phrase matching.
_SUBSIDIARY_FULL_NAMES = [
    ("south eastern coalfields", "SECL"),
    ("eastern coalfields", "ECL"),
    ("western coalfields", "WCL"),
    ("central coalfields", "CCL"),
    ("northern coalfields", "NCL"),
    ("mahanadi coalfields", "MCL"),
    ("bharat coking coal", "BCCL"),
    ("central mine planning", "CMPDI"),
]


def detect_subsidiary(text: str) -> Optional[str]:
    """Resolve a subsidiary code from an abbreviation or a spelled-out name."""
    if not text:
        return None
    m = SUBSIDIARY_RE.search(text)
    if m:
        return m.group(1).upper()
    low = text.lower()
    for phrase, code in _SUBSIDIARY_FULL_NAMES:
        if phrase in low:
            return code
    return None

# Canonical field for each recognised table-header keyword. Order matters:
# more specific metric columns are matched before the generic "mine" fallback,
# and "target" is checked before "production" so "Production Target MT" maps to
# the target column rather than actual production.
_HEADER_RULES = [
    ("subsidiary", ("subsidiary", "company")),
    ("report_year", ("year",)),
    ("production_target_mt", ("target", "budget", "planned")),
    ("actual_production_mt", ("actual", "achieved", "production", "output")),
    ("dispatch_mt", ("dispatch", "offtake")),
    ("overburden_mcum", ("overburden", "over burden")),
    ("mine_name", ("mine", "colliery", "project", "block")),
]

_EMPTY_RECORD = {
    "mine_name": None, "subsidiary": None, "report_year": None,
    "production_target_mt": None, "actual_production_mt": None,
    "dispatch_mt": None, "overburden_mcum": None, "unit": "MT",
}

_NUMERIC_FIELDS = (
    "production_target_mt", "actual_production_mt", "dispatch_mt", "overburden_mcum",
)


def _classify_header(cell: str) -> Optional[str]:
    """Map a single table-header cell to a canonical structured field, or None."""
    h = cell.strip().lower()
    if not h:
        return None
    # "OB" is the standard abbreviation for overburden; match it as a whole
    # token only (substring 'ob' would be too loose, e.g. "job", "problem").
    if h in ("ob", "o.b.", "o b") or re.fullmatch(r'ob\b.*', h):
        return "overburden_mcum"
    for field, keywords in _HEADER_RULES:
        if any(k in h for k in keywords):
            return field
    return None


def _to_number(cell: str) -> Optional[float]:
    """Parse the first numeric token out of a cell (tolerates units, commas)."""
    if cell is None:
        return None
    m = re.search(r'-?\d+(?:\.\d+)?', str(cell).replace(',', ''))
    return float(m.group(0)) if m else None


def _split_row(line: str) -> list[str]:
    """Split a parsed spreadsheet/table line on the '|' delimiter used upstream."""
    return [c.strip() for c in line.split('|')]


def _extract_table_records(text: str) -> list[dict]:
    """
    Detect a delimited header+data table (as produced by parse_spreadsheet /
    pipe-formatted OCR tables) and return one record per data row. Returns an
    empty list when no recognisable table is present.
    """
    records: list[dict] = []
    lines = [ln for ln in text.splitlines() if ln.strip()]
    header_map: Optional[dict] = None  # column index -> field
    ncols = 0

    for line in lines:
        # Skip sheet/page markers emitted by the parsers.
        if line.lstrip().startswith('===') or line.lstrip().startswith('---'):
            continue
        if '|' not in line:
            # A non-table line ends the current table block.
            header_map = None
            continue

        cells = _split_row(line)

        if header_map is None:
            mapping = {}
            for idx, cell in enumerate(cells):
                field = _classify_header(cell)
                if field and field not in mapping.values():
                    mapping[idx] = field
            # Require at least one metric column plus one more mapped column to
            # treat this as a genuine data table header (avoids false positives).
            has_metric = any(f in _NUMERIC_FIELDS for f in mapping.values())
            if len(mapping) >= 2 and has_metric:
                header_map = mapping
                ncols = len(cells)
            continue

        # We have a header; parse this row as data if the shape matches.
        if len(cells) != ncols:
            header_map = None
            continue

        rec = dict(_EMPTY_RECORD)
        for idx, field in header_map.items():
            raw = cells[idx] if idx < len(cells) else ""
            if field == "subsidiary":
                rec["subsidiary"] = detect_subsidiary(raw) or (raw.strip().upper() or None)
            elif field == "report_year":
                n = _to_number(raw)
                rec["report_year"] = int(n) if n is not None else None
            elif field == "mine_name":
                rec["mine_name"] = raw.strip()[:100] or None
            else:
                rec[field] = _to_number(raw)

        # Keep only rows that carry at least one real numeric metric.
        if any(rec[f] is not None for f in _NUMERIC_FIELDS):
            records.append(rec)

    return records


def _extract_prose_record(text: str) -> dict:
    """Single-record extraction for narrative documents via keyword regexes."""
    data = dict(_EMPTY_RECORD)

    data["subsidiary"] = detect_subsidiary(text)

    year_match = re.search(r'\b(20[12]\d)\b', text)
    if year_match:
        data["report_year"] = int(year_match.group(1))

    # Prefer the common "<Name> Mine/Colliery/OCP" pattern (e.g. "Kulda Mine"
    # -> "Kulda"); fall back to the "Mine: <Name>" label form.
    mine_match = re.search(
        r'\b([A-Z][A-Za-z0-9]+(?:\s+[A-Z][A-Za-z0-9]+){0,4})\s+(?:Open\s+Cast\s+Mine|OCP|Mine|Colliery|Coalfield|Project)\b',
        text,
    )
    if mine_match:
        data["mine_name"] = mine_match.group(1).strip()[:100]
    else:
        label_match = re.search(r'(?:mine|colliery|project)\s*[:\-]\s*([A-Za-z0-9][A-Za-z0-9\s]*?)(?=\,|\.|\n|$)', text, re.IGNORECASE)
        if label_match:
            data["mine_name"] = label_match.group(1).strip()[:100]

    # Connector allows "produced 45.6", "target of 47.0", "production was 45.6".
    _conn = r'\s*(?:of|:|-|=|was|at|stood\s+at)?\s*'
    prod_match = re.search(
        r'(?:actual\s+production|production|produced|achieved|raised|output)' + _conn + r'(\d+(?:\.\d+)?)\s*(?:MT|Million\s+Tonne)?',
        text, re.IGNORECASE)
    if prod_match:
        data["actual_production_mt"] = _to_number(prod_match.group(1))

    target_match = re.search(
        r'(?:target|planned|plan|budgeted)' + _conn + r'(\d+(?:\.\d+)?)\s*(?:MT|Million\s+Tonne)?',
        text, re.IGNORECASE)
    if target_match:
        data["production_target_mt"] = _to_number(target_match.group(1))

    dispatch_match = re.search(
        r'(?:dispatch(?:ed)?|despatch(?:ed)?|offtake)' + _conn + r'(\d+(?:\.\d+)?)\s*(?:MT|Million\s+Tonne)?',
        text, re.IGNORECASE)
    if dispatch_match:
        data["dispatch_mt"] = _to_number(dispatch_match.group(1))

    ob_match = re.search(r'(?:overburden|OB|removal)\s*[:\-]?\s*(\d+(?:\.\d+)?)\s*(?:MCuM|M3)?', text, re.IGNORECASE)
    if ob_match:
        data["overburden_mcum"] = _to_number(ob_match.group(1))

    return data


def extract_structured_records(text: str) -> list[dict]:
    """
    Return a list of structured mining records. Tabular documents yield one
    record per data row; narrative documents yield at most one prose-derived
    record. Empty list when nothing meaningful is found.
    """
    if not text:
        return []

    records = _extract_table_records(text)
    if records:
        # Propagate a document-level subsidiary to rows that lack their own,
        # so a report titled for a subsidiary still tags every mine row.
        doc_sub = detect_subsidiary(text)
        for rec in records:
            if not rec.get("subsidiary") and doc_sub:
                rec["subsidiary"] = doc_sub
        return records

    prose = _extract_prose_record(text)
    if any(prose[f] is not None for f in _NUMERIC_FIELDS) or prose["mine_name"]:
        return [prose]
    return []


def extract_structured_mining_data(text: str) -> dict:
    """Backward-compatible single-record view (first record, or empty stub)."""
    records = extract_structured_records(text)
    return records[0] if records else dict(_EMPTY_RECORD)


def download_file_from_s3(s3_key: str) -> bytes:
    """Helper to parse bucket & object key and download file bytes from MinIO."""
    s3_client = get_s3_client()
    key_clean = s3_key.lstrip('/')
    
    if '/' in key_clean:
        parts = key_clean.split('/', 1)
        # If first segment matches default bucket, separate it
        if parts[0] == MINIO_BUCKET:
            bucket_name = parts[0]
            object_name = parts[1]
        else:
            bucket_name = MINIO_BUCKET
            object_name = key_clean
    else:
        bucket_name = MINIO_BUCKET
        object_name = key_clean

    logger.info(f"Downloading s3 object: bucket={bucket_name}, key={object_name}")
    try:
        response = s3_client.get_object(Bucket=bucket_name, Key=object_name)
        return response['Body'].read()
    except Exception as e:
        logger.error(f"Failed to download {s3_key} from MinIO: {str(e)}")
        raise ValueError(f"MinIO download failed for s3_key '{s3_key}': {str(e)}")


def parse_pdf(file_bytes: bytes) -> tuple[str, str, int]:
    """
    Parses PDF document.
    First attempts direct text extraction with pypdf.
    If direct text is insufficient/empty, falls back to OCR via pdf2image + pytesseract.
    Returns (extracted_text, method_used, page_count).
    """
    page_count = 0
    # Step 1: Direct text extraction via pypdf
    try:
        pdf_reader = pypdf.PdfReader(io.BytesIO(file_bytes))
        page_count = len(pdf_reader.pages)
        direct_pages = []
        for i, page in enumerate(pdf_reader.pages):
            try:
                txt = page.extract_text() or ""
                if txt.strip():
                    direct_pages.append(f"--- Page {i+1} ---\n" + txt.strip())
            except Exception as pe:
                logger.warning(f"pypdf extraction error on page {i+1}: {str(pe)}")

        full_direct_text = "\n\n".join(direct_pages).strip()
        # If we found meaningful text (> 50 chars), use direct text extraction
        if len(full_direct_text) >= 50:
            logger.info("PDF text extraction succeeded directly via pypdf")
            return full_direct_text, "pdf_direct_text", page_count
    except Exception as e:
        logger.warning(f"Direct PDF text extraction failed, falling back to OCR: {str(e)}")

    # Step 2: Fallback to OCR via pdf2image and pytesseract
    def run_ocr_on_pdf():
        nonlocal page_count
        images = pdf2image.convert_from_bytes(file_bytes)
        page_count = len(images)
        ocr_pages = []
        for i, img in enumerate(images):
            try:
                page_text = pytesseract.image_to_string(img).strip()
                ocr_pages.append(f"--- Page {i+1} ---\n" + page_text)
            except Exception as pe:
                logger.error(f"OCR failed for page {i+1}: {str(pe)}")
                ocr_pages.append(f"--- Page {i+1} ---\n[OCR failed for this page: {str(pe)}]")
        return "\n\n".join(ocr_pages)

    executor = concurrent.futures.ThreadPoolExecutor(max_workers=1)
    future = executor.submit(run_ocr_on_pdf)
    try:
        ocr_text = future.result(timeout=OCR_TIMEOUT_SECONDS)
        return ocr_text, "pdf_ocr", page_count
    except concurrent.futures.TimeoutError:
        logger.error(f"PDF OCR timed out after {OCR_TIMEOUT_SECONDS}s")
        return (
            f"[Processing timed out after {OCR_TIMEOUT_SECONDS} seconds — needs manual review]",
            "pdf_ocr_timeout",
            page_count
        )
    finally:
        executor.shutdown(wait=False)


def parse_image(file_bytes: bytes) -> tuple[str, str, int]:
    """Parses image file using pytesseract OCR."""
    try:
        img = Image.open(io.BytesIO(file_bytes))
        ocr_text = pytesseract.image_to_string(img).strip()
        return ocr_text, "image_ocr", 1
    except Exception as e:
        raise ValueError(f"Image OCR processing failed: {str(e)}")


def parse_spreadsheet(file_bytes: bytes, s3_key: str) -> tuple[str, str, int]:
    """
    Parses spreadsheets (.xlsx, .xls, .csv, .tsv) into structured text representation.
    """
    ext = s3_key.lower().split('.')[-1] if '.' in s3_key else ''
    
    if ext in ['csv', 'tsv']:
        delimiter = '\t' if ext == 'tsv' else ','
        try:
            content_str = file_bytes.decode('utf-8', errors='ignore')
            reader = csv.reader(io.StringIO(content_str), delimiter=delimiter)
            rows = [ " | ".join(row) for row in reader if row ]
            text = "\n".join(rows)
            return text, "spreadsheet_direct_extraction", 1
        except Exception as e:
            raise ValueError(f"CSV/TSV parsing failed: {str(e)}")

    # Default to openpyxl for excel workbooks (.xlsx)
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        sheet_count = len(wb.sheetnames)
        sheet_texts = []
        for name in wb.sheetnames:
            try:
                sheet = wb[name]
                rows_text = []
                for row in sheet.iter_rows(values_only=True):
                    row_vals = [str(val) if val is not None else "" for val in row]
                    if any(row_vals):
                        rows_text.append(" | ".join(row_vals))
                if rows_text:
                    sheet_texts.append(f"=== Sheet: {name} ===\n" + "\n".join(rows_text))
            except Exception as se:
                logger.warning(f"Error reading sheet {name}: {str(se)}")
                sheet_texts.append(f"=== Sheet: {name} ===\n[Error reading sheet: {str(se)}]")

        full_text = "\n\n".join(sheet_texts).strip()
        return full_text, "spreadsheet_direct_extraction", sheet_count
    except Exception as e:
        raise ValueError(f"Spreadsheet parsing failed: {str(e)}")


_EVAL_FIELDS = [
    "mine_name", "subsidiary", "report_year",
    "production_target_mt", "actual_production_mt", "dispatch_mt", "overburden_mcum",
]
_BENCHMARK_PATH = os.path.join(os.path.dirname(__file__), "eval_benchmark.json")


def _norm_str(v) -> str:
    return re.sub(r'\s+', ' ', str(v).strip().lower()) if v is not None else ""


def _field_score(expected: dict, extracted: dict, tol: float) -> tuple[int, int, list]:
    """Count correctly-extracted fields for one matched record pair."""
    correct = 0
    total = 0
    wrong = []
    for f in _EVAL_FIELDS:
        if f not in expected or expected[f] is None:
            continue
        total += 1
        exp = expected[f]
        got = extracted.get(f)
        if isinstance(exp, (int, float)):
            ok = got is not None and abs(float(got) - float(exp)) <= tol
        else:
            ok = _norm_str(got) == _norm_str(exp)
        if ok:
            correct += 1
        else:
            wrong.append({"field": f, "expected": exp, "got": got})
    return correct, total, wrong


def _best_match(expected: dict, candidates: list, used: set, tol: float):
    """Greedily pick the unused extracted record that scores highest against expected."""
    best_i, best_correct, best = None, -1, (0, 0, [])
    for i, cand in enumerate(candidates):
        if i in used:
            continue
        c, t, w = _field_score(expected, cand, tol)
        if c > best_correct:
            best_i, best_correct, best = i, c, (c, t, w)
    return best_i, best


def run_extraction_eval() -> dict:
    """
    Run the REAL production extractor (extract_structured_records) against the
    labeled benchmark and compute field-level accuracy. No hardcoded numbers.
    """
    with open(_BENCHMARK_PATH, "r", encoding="utf-8") as fh:
        bench = json.load(fh)
    tol = float(bench.get("tolerance", 0.01))

    total_correct = 0
    total_fields = 0
    matched_records = 0
    expected_records = 0
    spurious_records = 0
    case_results = []

    for case in bench.get("cases", []):
        extracted = extract_structured_records(case["text"])
        expected = case["expected"]
        expected_records += len(expected)
        used = set()
        case_correct = case_total = 0
        case_wrong = []
        for exp in expected:
            i, (c, t, w) = _best_match(exp, extracted, used, tol)
            if i is not None:
                used.add(i)
                matched_records += 1
            case_correct += c
            case_total += t
            case_wrong.extend(w)
        spurious = len(extracted) - len(used)
        spurious_records += max(0, spurious)
        total_correct += case_correct
        total_fields += case_total
        case_results.append({
            "id": case["id"],
            "field_accuracy_pct": round(case_correct / case_total * 100.0, 2) if case_total else None,
            "fields_correct": case_correct,
            "fields_total": case_total,
            "records_expected": len(expected),
            "records_extracted": len(extracted),
            "spurious_records": max(0, spurious),
            "mismatches": case_wrong,
        })

    overall = round(total_correct / total_fields * 100.0, 2) if total_fields else None
    return {
        "evaluated": total_fields > 0,
        "dataset": bench.get("dataset"),
        "extraction_accuracy_percentage": overall,
        "fields_correct": total_correct,
        "fields_total": total_fields,
        "records_matched": matched_records,
        "records_expected": expected_records,
        "record_recall_pct": round(matched_records / expected_records * 100.0, 2) if expected_records else None,
        "spurious_records": spurious_records,
        "cases": case_results,
    }


@app.get("/health")
def health():
    return {"status": "ok"}


@app.get("/evaluate/extraction")
def evaluate_extraction():
    """Measured extraction accuracy against the labeled benchmark (live)."""
    try:
        return run_extraction_eval()
    except Exception as e:
        logger.error(f"Extraction evaluation failed: {str(e)}", exc_info=True)
        return {"evaluated": False, "extraction_accuracy_percentage": None, "error": str(e)}


@app.post("/extract", response_model=ExtractResponse)
def extract_text(payload: ExtractRequest):
    logger.info(f"Extract request received for document_id={payload.document_id}, s3_key={payload.s3_key}, source_type={payload.source_type}")
    
    # Download file from MinIO
    try:
        file_bytes = download_file_from_s3(payload.s3_key)
    except Exception as e:
        return ExtractResponse(
            document_id=payload.document_id,
            extracted_text="",
            method_used="failed",
            page_count=None,
            char_count=0,
            error=str(e)
        )

    # Route based on source_type or file extension
    source_type = payload.source_type.lower()
    ext = payload.s3_key.lower().split('.')[-1] if '.' in payload.s3_key else ''

    try:
        if source_type == 'pdf' or ext == 'pdf':
            text, method, pages = parse_pdf(file_bytes)
        elif source_type in ['image', 'png', 'jpg', 'jpeg'] or ext in ['png', 'jpg', 'jpeg', 'tiff', 'bmp', 'webp']:
            text, method, pages = parse_image(file_bytes)
        elif source_type in ['spreadsheet', 'excel', 'csv'] or ext in ['xlsx', 'xls', 'csv', 'tsv']:
            text, method, pages = parse_spreadsheet(file_bytes, payload.s3_key)
        else:
            # Fallback attempt based on content decoding or image loading
            try:
                text, method, pages = parse_image(file_bytes)
            except Exception:
                text = file_bytes.decode('utf-8', errors='ignore')
                method = "raw_text_decode"
                pages = 1

        struct_records = extract_structured_records(text)
        struct_data = struct_records[0] if struct_records else extract_structured_mining_data(text)
        logger.info(f"Structured extraction for {payload.document_id}: {len(struct_records)} record(s)")

        return ExtractResponse(
            document_id=payload.document_id,
            extracted_text=text,
            method_used=method,
            page_count=pages,
            char_count=len(text),
            structured_data=struct_data,
            structured_records=struct_records,
            error=None
        )
    except Exception as e:
        logger.error(f"Error processing extraction for {payload.document_id}: {str(e)}", exc_info=True)
        return ExtractResponse(
            document_id=payload.document_id,
            extracted_text="",
            method_used="failed",
            page_count=None,
            char_count=0,
            error=f"Extraction processing failed: {str(e)}"
        )
